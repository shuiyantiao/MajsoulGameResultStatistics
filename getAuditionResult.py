# statistics audition result
# summation last N consecutive games result as final result
# can filter result by min game count

# 统计海选赛结果
# 取最后连续N场比赛结果作为最终得分
# 可指定最小场次

import csv
import chardet
import openpyxl

filename = 'data.xlsx'
dataFileName = 'data.csv'

minGameCount = 1 # min game count for statistics 最小场次
lastContinuousGameCount = 3 # last N consecutive games count 计分的的最后连续场数

userDict = {}

def addData(name, score):
    if (name in userDict.keys()):
        userDict[name].append(score)
        pass
    else:
        userDict[name] = [score]
        pass
    pass

# get file encoding 获取文件编码类型
def get_encoding(file):
    with open(file, 'rb') as f:
        data = f.read()
        return chardet.detect(data)['encoding']

def intToChar(num):
    if 1 <= num <= 26:
        return chr(num + 64)
    return None

with open(dataFileName, 'rt', encoding=get_encoding(dataFileName)) as data:
    reader = csv.DictReader(data)
    for row in reader:
        addData(row['1位昵称'], float(row['1位得分']))
        addData(row['2位昵称'], float(row['2位得分']))
        addData(row['3位昵称'], float(row['3位得分']))
        addData(row['4位昵称'], float(row['4位得分']))

# get last N result 统计最后n场的分数
userScore = {}
maxLength = 0
for username, scores in userDict.items():
    length = len(scores)

    if length < minGameCount:
        continue 
    maxLength = length if length > maxLength else maxLength

    i = 0
    for j in range(length):
        if i == 0:
            userScore[username] = scores[j]
        else:
            userScore[username] += scores[j]
        i += 1
        if i == lastContinuousGameCount :
            break
    pass

scoreResults = sorted(userScore.items(), key=lambda x: x[1], reverse=True)

excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = '预选赛结果'

# 写入表头
sheet.cell(row=1, column=1, value='雀魂昵称')
for i in range(1, maxLength + 1):
    sheet.cell(row=1, column= i + 1, value='第' + str(i) + '场')
    pass
sheet.cell(row=1, column= maxLength + 1, value='总得分')
sheet.cell(row=1, column= maxLength + 2, value='排名')

# 写入数据
resultCount = len(scoreResults)
j = 2 # row 行号
for user, total in scoreResults:
    k = 1 # col 列号
    # nickname 昵称
    sheet.cell(row = j, column = k, value = user) 
    k += 1
    scores = userDict[user]
    scores.reverse()
    # games 比赛场次
    for score in scores:
        sheet.cell(row = j, column = k, value = score) 
        k += 1
        pass
    # last N result 最后n局得分
    k = maxLength + 1
    sheet.cell(row = j, column = maxLength + 1, value = total) 
    k += 1
    # final rank 总排名
    colSign = intToChar(k - 1)
    sheet.cell(row = j, column = k, value = f'=RANK({colSign}{j},${colSign}${2}:${colSign}${resultCount + 1})')

    j += 1
    pass

excel.save(filename)